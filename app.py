import streamlit as st
import sys, os, json, time, tempfile
sys.path.insert(0, '.')
from src.parsers.genesys_yaml_parser import GenesysYAMLParser
from src.agents.analyzer import IVRAnalyzer
from src.agents.documentor import IVRDocumentor

st.set_page_config(page_title='OrchestrIA', layout='wide', page_icon='🎙️',
                   initial_sidebar_state='collapsed')

st.html("""
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Mono:wght@300;400;500&family=Plus+Jakarta+Sans:wght@300;400;500;600&display=swap" rel="stylesheet">
<style>
*, *::before, *::after { box-sizing: border-box; }
html, body, [data-testid="stApp"] {
    background: #07080B !important; color: #E8EDF5 !important;
    font-family: 'Plus Jakarta Sans', sans-serif !important; }
[data-testid="stAppViewContainer"] > .main { background: #07080B !important; }
.block-container { padding: 2rem 2.5rem 4rem !important; max-width: 1320px !important; }
#MainMenu, footer, header, [data-testid="stToolbar"],
[data-testid="stDecoration"], [data-testid="stStatusWidget"] { display: none !important; }
h1,h2,h3 { font-family: 'Syne', sans-serif !important; }
hr { border-color: #1C2030 !important; margin: 1.75rem 0 !important; }

/* HEADER */
.orch-header { display:flex; align-items:center; justify-content:space-between;
    padding:0 0 1.75rem; border-bottom:1px solid #1C2030; margin-bottom:2rem; }
.orch-logo-main { font-family:'Syne',sans-serif; font-size:1.5rem; font-weight:800;
    color:#F0F6FC; letter-spacing:-0.03em; }
.orch-logo-dot { color:#00D4AA; font-size:1.5rem; font-weight:800; }
.orch-logo-badge { font-family:'DM Mono',monospace; font-size:0.68rem; color:#3D4D66;
    background:#0E1118; border:1px solid #1C2030; border-radius:4px;
    padding:2px 8px; margin-left:0.6rem; letter-spacing:0.1em; text-transform:uppercase; }
.orch-tagline { font-family:'DM Mono',monospace; font-size:0.68rem; color:#3D4D66;
    letter-spacing:0.1em; text-transform:uppercase; }

/* RADIO TABS */
.stRadio > div { flex-direction:row !important; gap:0 !important;
    background:#0E1118 !important; border:1px solid #1C2030 !important;
    border-radius:8px !important; padding:3px !important;
    display:inline-flex !important; width:auto !important; }
.stRadio > div > label { background:transparent !important; border:none !important;
    border-radius:5px !important; padding:7px 18px !important;
    font-family:'Plus Jakarta Sans',sans-serif !important; font-size:0.82rem !important;
    font-weight:500 !important; color:#4B5568 !important;
    cursor:pointer !important; transition:all 0.15s !important; white-space:nowrap !important; }
.stRadio > div > label:has(input:checked) { background:#1C2030 !important; color:#E8EDF5 !important; }
.stRadio > div > label > div:first-child { display:none !important; }

/* UPLOAD */
[data-testid="stFileUploader"] { background:#0B0D14 !important;
    border:1px dashed #1C2030 !important; border-radius:10px !important; padding:0.25rem !important; }
[data-testid="stFileUploaderDropzone"] { background:transparent !important;
    border:none !important; padding:1.5rem !important; }
[data-testid="stFileUploaderDropzoneInstructions"] p {
    font-family:'Plus Jakarta Sans',sans-serif !important; color:#4B5568 !important; font-size:0.82rem !important; }

/* TEXTAREA */
.stTextArea textarea { background:#0B0D14 !important; border:1px solid #1C2030 !important;
    border-radius:10px !important; color:#8B9EB8 !important;
    font-family:'DM Mono',monospace !important; font-size:0.76rem !important; line-height:1.65 !important; }
.stTextArea textarea:focus { border-color:#00D4AA40 !important; outline:none !important; }

/* BUTTONS */
.stButton > button { background:#00D4AA !important; color:#07080B !important;
    border:none !important; border-radius:8px !important;
    font-family:'Plus Jakarta Sans',sans-serif !important; font-size:0.85rem !important;
    font-weight:600 !important; padding:0.65rem 1.5rem !important; transition:all 0.15s !important; }
.stButton > button:hover { background:#00BBAA !important; transform:translateY(-1px) !important;
    box-shadow:0 4px 20px #00D4AA25 !important; }
.stDownloadButton > button { background:#0E1118 !important; color:#00D4AA !important;
    border:1px solid #00D4AA40 !important; border-radius:8px !important;
    font-family:'Plus Jakarta Sans',sans-serif !important; font-size:0.82rem !important; }
.stDownloadButton > button:hover { background:#00D4AA12 !important; }

/* METRICS */
div[data-testid="metric-container"] { background:#0B0D14 !important;
    border:1px solid #1C2030 !important; border-radius:10px !important; padding:1rem 1.1rem !important; }
div[data-testid="metric-container"] label { font-family:'DM Mono',monospace !important;
    font-size:0.62rem !important; color:#3D4D66 !important;
    text-transform:uppercase !important; letter-spacing:0.12em !important; }
div[data-testid="metric-container"] [data-testid="stMetricValue"] {
    font-family:'Syne',sans-serif !important; font-size:1.5rem !important;
    font-weight:700 !important; color:#E8EDF5 !important; }

/* PROGRESS */
.stProgress > div > div { background:#0E1118 !important; border-radius:4px !important; }
.stProgress > div > div > div { background:linear-gradient(90deg,#00D4AA,#00A8FF) !important; border-radius:4px !important; }

/* EXPANDER */
.streamlit-expanderHeader { background:#0B0D14 !important; border:1px solid #1C2030 !important;
    border-radius:8px !important; font-family:'Plus Jakarta Sans',sans-serif !important;
    font-size:0.85rem !important; color:#8B9EB8 !important; padding:0.75rem 1rem !important; }
.streamlit-expanderContent { background:#09090F !important; border:1px solid #1C2030 !important;
    border-top:none !important; border-radius:0 0 8px 8px !important; padding:1.5rem !important; }

/* CUSTOM CLASSES */
.orch-badge { display:inline-flex; align-items:center; gap:6px; padding:3px 10px;
    border-radius:20px; font-family:'DM Mono',monospace; font-size:0.7rem;
    font-weight:500; letter-spacing:0.05em; }
.badge-simple   { background:#00D4AA12; color:#00D4AA; border:1px solid #00D4AA25; }
.badge-moderado { background:#D2992212; color:#D29922; border:1px solid #D2992225; }
.badge-complejo { background:#F0883E12; color:#F0883E; border:1px solid #F0883E25; }
.badge-muy      { background:#F8514912; color:#F85149; border:1px solid #F8514925; }

.orch-chip { display:inline-flex; align-items:center; background:#0E1118;
    border:1px solid #1C2030; border-radius:5px; padding:2px 9px;
    font-family:'DM Mono',monospace; font-size:0.7rem; color:#6B7A94; margin:2px; }
.chip-teal { color:#00D4AA; border-color:#00D4AA20; background:#00D4AA06; }
.chip-blue { color:#00A8FF; border-color:#00A8FF20; background:#00A8FF06; }

.lbl { font-family:'DM Mono',monospace; font-size:0.62rem; color:#3D4D66;
    text-transform:uppercase; letter-spacing:0.15em; margin-bottom:0.85rem; display:block; }

.ph-done  { color:#00D4AA; font-family:'DM Mono',monospace; font-size:0.78rem; display:block; padding:7px 0; border-bottom:1px solid #13161E; }
.ph-active{ color:#00A8FF; font-family:'DM Mono',monospace; font-size:0.78rem; display:block; padding:7px 0; border-bottom:1px solid #13161E; }
.ph-pend  { color:#252D3D; font-family:'DM Mono',monospace; font-size:0.78rem; display:block; padding:7px 0; border-bottom:1px solid #13161E; }

/* FEATURE CARDS (empty state) */
.feat-card { background:#0B0D14; border:1px solid #1C2030; border-radius:10px;
    padding:1.25rem 1.5rem; margin-bottom:0.75rem; }
.feat-icon { font-size:1.1rem; margin-bottom:0.5rem; display:block; }
.feat-title { font-family:'Syne',sans-serif; font-size:0.88rem; font-weight:700;
    color:#E8EDF5; margin-bottom:0.35rem; }
.feat-desc { font-family:'Plus Jakarta Sans',sans-serif; font-size:0.8rem;
    color:#4B5568; line-height:1.55; }

/* GENESYS COMPAT CHIPS */
.compat-row { display:flex; flex-wrap:wrap; gap:6px; margin-top:0.6rem; }
.compat-chip { font-family:'DM Mono',monospace; font-size:0.65rem;
    background:#0E1118; border:1px solid #1C2030; border-radius:4px;
    padding:2px 8px; color:#3D4D66; }
</style>
""")

# ── SESSION STATE ──────────────────────────────────────────────────────────────
for k, v in [('analysis', None), ('flow', None),
              ('batch_results', []), ('batch_flows', {})]:
    if k not in st.session_state:
        st.session_state[k] = v

# ── HELPERS ────────────────────────────────────────────────────────────────────
def parse_content(content, filename):
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else 'yaml'
    if ext == 'json':
        try:
            import yaml; content = yaml.dump(json.loads(content))
        except Exception as e: return None, 'Error JSON: ' + str(e)
    elif ext == 'xml':
        try:
            import xmltodict, yaml; content = yaml.dump(dict(xmltodict.parse(content)))
        except ImportError: return None, 'xmltodict no instalado: pip install xmltodict'
        except Exception as e: return None, 'Error XML: ' + str(e)
    return GenesysYAMLParser().parse(content, flow_name=filename.rsplit('.', 1)[0]), None

def generar_pdf_bytes(flow, analysis):
    # Enriquecer el analysis con drivers textuales del score
    enriched = dict(analysis)
    inv = enriched.get('inventory', {})
    bd  = inv.get('migration_score_breakdown', {})
    ml  = inv.get('migration_level', '')
    
    # Añadir drivers de migración como risk flags extra
    mig_drivers = []
    if inv.get('data_services'):
        mig_drivers.append(f"APIs de datos: {', '.join(inv['data_services'])} — reconexión requerida en Cloud")
    if inv.get('auth_services'):
        mig_drivers.append(f"Auth services: {', '.join(inv['auth_services'])} — validar OAuth/SAML en Cloud")
    if inv.get('dynamic_variables'):
        mig_drivers.append(f"{len(inv['dynamic_variables'])} variable(s) TTS dinámica(s) — verificar runtime Cloud")
    if not inv.get('entry_node_id'):
        mig_drivers.append("Entry node no definido — prerequisito arquitectónico de la migración")
    dead = inv.get('dead_ends', [])
    if dead:
        mig_drivers.append(f"{len(dead)} dead end(s) — requieren rediseño antes de migrar")
    
    # Reemplazar flags con versión enriquecida (sin duplicar)
    # Los drivers textuales son más informativos que los flags originales
    if 'inventory' in enriched and mig_drivers:
        enriched['inventory'] = dict(enriched['inventory'])
        enriched['inventory']['migration_risk_flags'] = mig_drivers

    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp: path = tmp.name
    IVRDocumentor().generate_pdf(flow, enriched, path)
    with open(path, 'rb') as f: data = f.read()
    os.unlink(path); return data

def score_color(s):
    return '#00D4AA' if s >= 70 else '#D29922' if s >= 40 else '#F85149'

def score_ring(score):
    color = score_color(score)
    r, size = 52, 124
    circ = 2 * 3.14159 * r
    offset = circ * (1 - score / 100)
    return (
        f'<div style="display:flex;flex-direction:column;align-items:center;padding:0.75rem 0 1.25rem;">'
        f'<div style="font-family:\'DM Mono\',monospace;font-size:0.58rem;color:#3D4D66;'
        f'text-transform:uppercase;letter-spacing:0.2em;margin-bottom:0.4rem;">Quality Score</div>'
        f'<div style="position:relative;width:{size}px;height:{size}px;">'
        f'<svg width="{size}" height="{size}" style="transform:rotate(-90deg)">'
        f'<circle cx="{size//2}" cy="{size//2}" r="{r}" fill="none" stroke="#13161E" stroke-width="5"/>'
        f'<circle cx="{size//2}" cy="{size//2}" r="{r}" fill="none" stroke="{color}"'
        f' stroke-width="5" stroke-linecap="round"'
        f' stroke-dasharray="{circ:.1f}" stroke-dashoffset="{offset:.1f}"'
        f' style="filter:drop-shadow(0 0 7px {color}55)"/>'
        f'</svg>'
        f'<div style="position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);text-align:center;">'
        f'<div style="font-family:\'Syne\',sans-serif;font-size:2.2rem;font-weight:800;'
        f'color:{color};line-height:1;letter-spacing:-0.04em;">{score}</div>'
        f'</div></div>'
        f'<div style="font-family:\'DM Mono\',monospace;font-size:0.7rem;color:#3D4D66;margin-top:0.15rem;">/ 100</div>'
        f'</div>'
    )

def migration_badge(ml):
    cls = {'SIMPLE':'badge-simple','MODERADO':'badge-moderado',
           'COMPLEJO':'badge-complejo','MUY COMPLEJO':'badge-muy'}.get(ml,'badge-simple')
    return f'<span class="orch-badge {cls}">● {ml}</span>'

def empty_state_panel():
    return """
<div style="padding:0.25rem 0 0;">
  <div style="font-family:'DM Mono',monospace;font-size:0.58rem;color:#252D3D;
       text-transform:uppercase;letter-spacing:0.15em;margin-bottom:1.25rem;">What you'll get</div>

  <div class="feat-card">
    <span class="feat-icon">◈</span>
    <div class="feat-title">Quality Score + Rubric</div>
    <div class="feat-desc">AI-powered audit across 6 dimensions — structural integrity, operational robustness, CX, API handling, architecture, and edge coverage.</div>
  </div>

  <div class="feat-card">
    <span class="feat-icon">⬡</span>
    <div class="feat-title">Full Flow Inventory</div>
    <div class="feat-desc">Node counts, queue mapping, external API dependencies, dynamic TTS variables, auth services, self-service ratio.</div>
  </div>

  <div class="feat-card">
    <span class="feat-icon">⇢</span>
    <div class="feat-title">Migration Assessment</div>
    <div class="feat-desc">Complexity scoring for Genesys Cloud migration. Risk flags, dependency matrix, estimated effort level.</div>
  </div>

  <div class="feat-card">
    <span class="feat-icon">↓</span>
    <div class="feat-title">Executive PDF Report</div>
    <div class="feat-desc">CIO-ready PDF with findings, action plan, migration roadmap, and technical inventory. Ready to share in minutes.</div>
  </div>

  <div style="margin-top:1.25rem;">
    <div style="font-family:'DM Mono',monospace;font-size:0.58rem;color:#252D3D;
         text-transform:uppercase;letter-spacing:0.15em;margin-bottom:0.6rem;">Genesys compatibility</div>
    <div class="compat-row">
      <span class="compat-chip">Genesys Cloud CX</span>
      <span class="compat-chip">Architect inboundCall</span>
      <span class="compat-chip">YAML · JSON · XML</span>
      <span class="compat-chip">dataQuery / apiCall</span>
      <span class="compat-chip">authenticate</span>
      <span class="compat-chip">TTS Variables</span>
      <span class="compat-chip">Transfer Nodes</span>
    </div>
  </div>
</div>"""

def mostrar_loading(placeholder):
    fases = ['Parsing flow structure','Extracting node inventory',
             'Detecting external dependencies','AI analysis · Genesys Expert',
             'Generating migration assessment']
    hints = ['Scanning for dead ends...','Mapping transfer targets...',
             'Evaluating timeout coverage...','Cross-referencing node graph...',
             'Calculating migration complexity...']
    def render(step, hi):
        rows = ''.join(
            f'<span class="ph-{"done" if i<step else "active" if i==step else "pend"}">'
            f'{"✓" if i<step else "▸" if i==step else "·"} {label}</span>'
            for i, label in enumerate(fases))
        with placeholder.container():
            st.markdown(
                f'<div style="background:#0B0D14;border:1px solid #1C2030;border-radius:10px;'
                f'padding:1.25rem 1.5rem;"><span class="lbl">Analyzing</span>'
                f'{rows}'
                f'<div style="margin-top:0.85rem;font-family:\'DM Mono\',monospace;'
                f'font-size:0.65rem;color:#252D3D;">{hints[hi%len(hints)]}</div></div>',
                unsafe_allow_html=True)
            st.progress((step+1)/len(fases))
    return render

def mostrar_resultado(analysis, flow=None, key_prefix='main'):
    score  = analysis.get('score', 0)
    inv    = analysis.get('inventory', {})
    issues = analysis.get('critical_issues', [])
    imps   = analysis.get('improvements', [])
    summ   = analysis.get('summary', '')

    col_s, col_r = st.columns([1, 2])
    with col_s:
        st.markdown(score_ring(score), unsafe_allow_html=True)
        # Export buttons rendered by render_export_buttons() above

    with col_r:
        st.markdown(
            f'<div style="font-family:\'Plus Jakarta Sans\',sans-serif;font-size:0.9rem;'
            f'color:#7A8BA5;line-height:1.75;padding:0.25rem 0 1rem;">{summ}</div>',
            unsafe_allow_html=True)
        for i in issues: st.error(i)
        for i in imps:   st.success(i)

    if inv:
        st.divider()
        st.markdown('<span class="lbl">Flow Inventory</span>', unsafe_allow_html=True)
        c1,c2,c3,c4,c5,c6 = st.columns(6)
        c1.metric('Nodes',    inv.get('total_nodes',0))
        c2.metric('Menus',    inv.get('menu_nodes',0))
        c3.metric('Transfers',inv.get('transfer_nodes',0))
        c4.metric('Logic',    inv.get('task_nodes',0))
        c5.metric('Self-Svc',str(inv.get('self_service_ratio',0))+'%')
        c6.metric('Ext. Deps',inv.get('total_external_deps',0))

    if inv and any([inv.get('data_services'),inv.get('auth_services'),
                    inv.get('dynamic_variables'),inv.get('unique_queues')]):
        st.divider()
        st.markdown('<span class="lbl">External Dependencies</span>', unsafe_allow_html=True)
        chips = ''
        for s in inv.get('data_services',   []): chips += f'<span class="orch-chip chip-teal">⬡ {s}</span>'
        for s in inv.get('auth_services',   []): chips += f'<span class="orch-chip chip-blue">⬡ {s}</span>'
        for v in inv.get('dynamic_variables',[]): chips += f'<span class="orch-chip">'+'{'+v+'}</span>'
        for q in inv.get('unique_queues',   []): chips += f'<span class="orch-chip">⇒ {q}</span>'
        st.markdown('<div style="display:flex;flex-wrap:wrap;gap:0.2rem;">'+chips+'</div>',
                    unsafe_allow_html=True)

    if inv:
        st.divider()
        ml = inv.get('migration_level', 'SIMPLE')
        ms = inv.get('migration_complexity_score', 0)
        breakdown = inv.get('migration_score_breakdown', {})
        flags = inv.get('migration_risk_flags', [])

        # Header del card — benchmark fijo banca
        st.markdown(
            f'<div style="display:flex;align-items:center;gap:0.9rem;margin-bottom:0.5rem;">'            f'<span class="lbl" style="margin:0;">Migration to Cloud</span>'            f'{migration_badge(ml)}'            f'<span style="font-family:\'DM Mono\',monospace;font-size:0.7rem;color:#3D4D66;">{ms}/100</span>'            f'</div>', unsafe_allow_html=True)
        st.markdown(benchmark_card(inv, 'banking'), unsafe_allow_html=True)


        # Breakdown de 5 dimensiones si existe
        if breakdown:
            dim_colors = {
                'D1_grafo':        '#00A8FF',
                'D2_dependencias': '#00D4AA',
                'D3_riesgo':       '#F85149',
                'D4_escala':       '#D29922',
                'D5_testing':      '#A78BFA',
            }
            bars_html = '<div style="display:flex;flex-direction:column;gap:8px;margin-bottom:1.25rem;">'
            for key, dim in breakdown.items():
                label  = dim.get('label', key)
                dscore = dim.get('score', 0)
                dmax   = dim.get('max', 25)
                pct    = round(dscore / dmax * 100) if dmax else 0
                color  = dim_colors.get(key, '#4B5568')
                bars_html += (
                    f'<div style="display:flex;align-items:center;gap:10px;">'
                    f'<div style="font-family:\'DM Mono\',monospace;font-size:0.62rem;'
                    f'color:#4B5568;width:160px;flex-shrink:0;">{label}</div>'
                    f'<div style="flex:1;background:#0E1118;border-radius:3px;height:5px;overflow:hidden;">'
                    f'<div style="width:{pct}%;background:{color};height:100%;border-radius:3px;'
                    f'box-shadow:0 0 6px {color}50;"></div></div>'
                    f'<div style="font-family:\'DM Mono\',monospace;font-size:0.62rem;'
                    f'color:#3D4D66;width:40px;text-align:right;">{dscore}/{dmax}</div>'
                    f'</div>'
                )
            bars_html += '</div>'
            st.markdown(bars_html, unsafe_allow_html=True)

        # Score & Migration explanation — determinista, sin LLM
        explanation = score_explanation(analysis)
        if explanation:
            st.markdown(explanation, unsafe_allow_html=True)

        # Migration hours estimate — sección destacada
        st.markdown(migration_hours_card(inv), unsafe_allow_html=True)

        # Risk flags
        if flags:
            for f_ in flags:
                st.warning(f_)
        else:
            st.success('No migration risks detected')



from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, re
from datetime import date

# ── PALETTE ───────────────────────────────────────────────────────────────────
BG="07080B"; SURFACE="0E1118"; CARD="161B22"; BORDER="1C2030"
TEXT="E8EDF5"; DIM="4B5568"; WHITE="FFFFFF"
TEAL="00D4AA"; RED="F85149"; YELLOW="D29922"; ORANGE="F0883E"
BLUE="00A8FF"; PURPLE="A78BFA"; GREEN="3FB950"

def hf(c): return PatternFill("solid", fgColor=c)
def bb(c=BORDER): return Border(bottom=Side(style="thin", color=c))
def tb(c=BORDER): return Border(top=Side(style="thin", color=c))
def full_border(c=BORDER):
    s=Side(style="thin",color=c)
    return Border(left=s,right=s,top=s,bottom=s)
def score_color(s): return TEAL if s>=70 else YELLOW if s>=40 else RED
def mig_color(l): return {"SIMPLE":TEAL,"MODERADO":YELLOW,"COMPLEJO":ORANGE,"MUY COMPLEJO":RED}.get(l,"4B5568")
def safe_name(s): return re.sub(r'[\\/*?:\[\]]','_',s)[:31]

def set_cols(ws, widths):
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w

def bg_all(ws, rows=400, cols=20):
    for row in ws.iter_rows(min_row=1,max_row=rows,min_col=1,max_col=cols):
        for c in row: c.fill=hf(BG)

def section_header(ws, row, col_start, col_end, label, color=TEAL):
    ws.merge_cells(f"{get_column_letter(col_start)}{row}:{get_column_letter(col_end)}{row}")
    c = ws.cell(row=row, column=col_start, value=label)
    c.font = Font(name="Arial", bold=True, size=8, color=color)
    c.fill = hf(SURFACE)
    c.alignment = Alignment(vertical="center")
    c.border = bb()
    ws.row_dimensions[row].height = 18
    return row+1

def table_header(ws, row, headers, col_start=2):
    ws.row_dimensions[row].height = 18
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start+i, value=h)
        c.font = Font(name="Arial", bold=True, size=8, color=DIM)
        c.fill = hf(SURFACE)
        c.alignment = Alignment(horizontal="center" if i>0 else "left", vertical="center")
        c.border = bb()
    return row+1

def data_row(ws, row, values, col_start=2, alt=False, colors=None):
    ws.row_dimensions[row].height = 17
    rbg = CARD if alt else SURFACE
    for i, val in enumerate(values):
        col = col_start+i
        c = ws.cell(row=row, column=col, value=val)
        c.fill = hf(rbg)
        c.border = bb(BORDER)
        txt_color = colors[i] if colors and i < len(colors) else TEXT
        c.font = Font(name="Arial", size=9, color=txt_color)
        c.alignment = Alignment(
            horizontal="left" if i==0 else "center",
            vertical="center", wrap_text=(i==0)
        )
    return row+1

def spacer(ws, row, height=8):
    ws.row_dimensions[row].height = height
    return row+1


# ── SHEET BUILDER: PORTFOLIO OVERVIEW ─────────────────────────────────────────
def build_overview(wb, data, today):
    ws = wb.active; ws.title="Portfolio Overview"
    ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor=TEAL
    ws.freeze_panes="A6"; bg_all(ws)
    set_cols(ws,[3,30,8,12,8,8,10,12,10,14,12,12])
    ws.row_dimensions[1].height=6; ws.row_dimensions[2].height=38
    ws.row_dimensions[3].height=6; ws.row_dimensions[4].height=16; ws.row_dimensions[5].height=20

    ws["B2"].value="OrchestrIA"
    ws["B2"].font=Font(name="Arial",bold=True,size=18,color=WHITE)
    ws["B2"].alignment=Alignment(vertical="center")
    ws["D2"].value="IVR · IA"
    ws["D2"].font=Font(name="Arial",size=11,color=TEAL)
    ws["D2"].alignment=Alignment(vertical="center")
    ws["H2"].value=f"Portfolio Analysis  ·  {len(data)} flows  ·  {today}"
    ws["H2"].font=Font(name="Arial",size=8,color=DIM)
    ws["H2"].alignment=Alignment(vertical="center",horizontal="right")

    ws["B4"].value="Flows ranked by Quality Score. Click sheet tabs for individual flow detail."
    ws["B4"].font=Font(name="Arial",size=8,color=DIM,italic=True)
    ws["B4"].alignment=Alignment(vertical="center")

    hdrs=["","Flow Name","Score","Quality","Nodes","Menus","Xfer","Self-Svc%","Ext Deps","Migration","Mig.Score","Dead Ends"]
    for col,h in enumerate(hdrs,1):
        c=ws.cell(row=5,column=col,value=h)
        c.font=Font(name="Arial",bold=True,size=8,color=DIM)
        c.fill=hf(SURFACE); c.border=bb()
        c.alignment=Alignment(horizontal="left" if col<=2 else "center",vertical="center")

    for idx,r in enumerate(data):
        row=6+idx; ws.row_dimensions[row].height=19
        inv=r.get("inventory",{}); score=r.get("score",0)
        ml=inv.get("migration_level","—"); ms=inv.get("migration_complexity_score",0)
        fname=r["filename"].replace(".yaml","").replace(".yml","")
        rbg=CARD if idx%2==0 else SURFACE
        if score>=70:   ql,qc="● GOOD",TEAL
        elif score>=40: ql,qc="● FAIR",YELLOW
        else:           ql,qc="● POOR",RED
        vals=["",fname,score,ql,inv.get("total_nodes",0),inv.get("menu_nodes",0),
              inv.get("transfer_nodes",0),inv.get("self_service_ratio",0),
              inv.get("total_external_deps",0),ml,ms,len(inv.get("dead_ends",[]))]
        for col,val in enumerate(vals,1):
            c=ws.cell(row=row,column=col,value=val)
            c.fill=hf(rbg); c.border=bb(BORDER)
            c.alignment=Alignment(horizontal="left" if col<=2 else "center",vertical="center")
            c.font=Font(name="Arial",size=9,color=TEXT)
        ws.cell(row=row,column=3).font=Font(name="Arial",bold=True,size=11,color=score_color(score))
        ws.cell(row=row,column=4).font=Font(name="Arial",size=8,bold=True,color=qc)
        ws.cell(row=row,column=10).font=Font(name="Arial",size=8,bold=True,color=mig_color(ml))
        ws.cell(row=row,column=8).number_format="0.0"

    sum_row=6+len(data)+1; ws.row_dimensions[sum_row].height=24
    ws.cell(row=sum_row,column=2,value="Portfolio Average")
    ws.cell(row=sum_row,column=2).font=Font(name="Arial",bold=True,size=9,color=TEAL)
    ws.cell(row=sum_row,column=2).fill=hf(SURFACE)
    avg=ws.cell(row=sum_row,column=3)
    avg.value=f"=AVERAGE(C6:C{5+len(data)})"
    avg.font=Font(name="Arial",bold=True,size=12,color=TEAL)
    avg.fill=hf(SURFACE); avg.number_format="0"
    avg.alignment=Alignment(horizontal="center",vertical="center")
    for col in [1]+list(range(4,13)): ws.cell(row=sum_row,column=col).fill=hf(SURFACE)


# ── SHEET BUILDER: INDIVIDUAL FLOW ────────────────────────────────────────────
def build_flow_sheet(wb, r, raw_yaml=None):
    inv   = r.get("inventory", {})
    score = r.get("score", 0)
    fname = r["filename"].replace(".yaml","").replace(".yml","")
    tab_name = safe_name(fname[:31])

    ws = wb.create_sheet(tab_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = score_color(score)
    bg_all(ws)
    set_cols(ws,[3,22,18,12,26,12,12,12,16])
    ws.freeze_panes = "A7"

    # ── HEADER ────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height=6; ws.row_dimensions[2].height=36
    ws.row_dimensions[3].height=20; ws.row_dimensions[4].height=6

    ws["B2"].value=fname
    ws["B2"].font=Font(name="Arial",bold=True,size=15,color=WHITE)
    ws["B2"].alignment=Alignment(vertical="center")

    ml=inv.get("migration_level","—"); ms=inv.get("migration_complexity_score",0)
    ws["F2"].value=f"Score  {score}/100"
    ws["F2"].font=Font(name="Arial",bold=True,size=13,color=score_color(score))
    ws["F2"].alignment=Alignment(vertical="center")

    ws["H2"].value=f"Migration: {ml}  ({ms}/100)"
    ws["H2"].font=Font(name="Arial",size=10,bold=True,color=mig_color(ml))
    ws["H2"].alignment=Alignment(vertical="center",horizontal="right")

    # Summary
    summary = r.get("summary","") or r.get("executive_summary","")
    if summary:
        ws["B3"].value=summary[:200]
        ws["B3"].font=Font(name="Arial",size=8,color=DIM,italic=True)
        ws["B3"].alignment=Alignment(vertical="center",wrap_text=True)
        ws.merge_cells("B3:I3")
    ws.row_dimensions[3].height=32

    # ── S1: NODE INVENTORY ────────────────────────────────────────────────────
    row=5
    row=section_header(ws,row,2,9,"§1 · NODE INVENTORY",TEAL)
    row=table_header(ws,row,["Node ID","Name","Type","Next Nodes","TTS Prompt","Timeout","MaxRetries","Dead End?","Est. Hours"])
    nodes = inv.get("nodes_detail") or []
    if not nodes and raw_yaml:
        # Extraer nodos directamente del YAML raw
        SECTION_TYPE = {
            "menus":"MENU","tasks":"TASK","transfers":"TRANSFER",
            "prompts":"PROMPT","tasks_voicemail":"VOICEMAIL","exits":"EXIT"
        }
        dead_ends = set(inv.get("dead_ends",[]))
        for section, ntype in SECTION_TYPE.items():
            for nid, ndata in (raw_yaml.get(section) or {}).items():
                if not isinstance(ndata,dict): continue
                # next nodes
                nexts=[]
                for ch in (ndata.get("choices") or []):
                    if isinstance(ch,dict) and ch.get("next"):
                        nexts.append(ch["next"].split("/")[-1])
                for k in ("next","onSuccess","onFailure","onTimeout","onMaxRetries","fallback"):
                    v=ndata.get(k)
                    if isinstance(v,dict) and v.get("next"): nexts.append(v["next"].split("/")[-1])
                    elif isinstance(v,str) and v.startswith("./"): nexts.append(v.split("/")[-1])
                # TTS
                tts = ndata.get("tts","") or ndata.get("initialGreeting",{}).get("tts","") if isinstance(ndata.get("initialGreeting"),dict) else ""
                if isinstance(tts,str): tts=tts[:60]+"…" if len(tts)>60 else tts
                timeout = ndata.get("timeout","")
                maxret  = ndata.get("maxRetries","")
                is_dead = "⚠ YES" if nid in dead_ends else ""
                nodes.append({
                    "id":nid,"name":ndata.get("name",nid),"type":ntype,
                    "next":", ".join(nexts[:4]) or "—","tts":tts or "—",
                    "timeout":timeout or "—","maxRetries":maxret or "—",
                    "dead_end":is_dead
                })

    total_hours = 0
    if nodes:
        for idx,n in enumerate(nodes):
            if isinstance(n,dict):
                vals=[n.get("id",n.get("node_id","?")),n.get("name",""),n.get("type",""),
                      n.get("next","—"),n.get("tts","—"),n.get("timeout","—"),
                      n.get("maxRetries","—"),n.get("dead_end","")]
            else:
                vals=[str(n),"","","","","","",""]
            # ── EST. HOURS — Metodología integrador certificado Genesys ──────────
            # Basado en Cyara Cloud Migration Assurance + TTEC Digital + PS Genesys
            # No son datos oficiales Genesys — ver disclaimer en footer del Excel
            HOURS_BY_TYPE = {
                "ENTRY":     0.5,  # Config DNIS + inicio de flujo
                "MENU":      3.0,  # DTMF options + prompts + test (sin timeout = +1h)
                "PROMPT":    1.0,  # Audio upload + TTS config
                "SPEECH":    3.0,  # ASR + NLU mapping
                "INPUT":     2.0,  # collectInput + validación
                "CONDITION": 1.5,  # Lógica booleana + variables
                "SWITCH":    1.5,  # Multi-branch logic
                "SET_VARIABLE": 1.0, # Mapeo de variables
                "LOOP":      2.0,  # Lógica iteración + test bucle
                "CALLBACK":  4.0,  # Queue + callback scheduling
                "TRANSFER":  4.0,  # Queue mapping + skills + horarios
                "TASK":      3.0,  # Lógica interna sin integración
                "VOICEMAIL": 3.0,  # Voicemail config + transcripción
                "EXIT":      0.5,  # Disconnect action
                "UNKNOWN":   2.0,  # Requiere revisión manual
            }
            ntype_upper = str(vals[2]).upper()
            est_h = HOURS_BY_TYPE.get(ntype_upper, 2.0)
            # Multiplicador: si el nodo tiene API/data dip → 8h (Data Action definition)
            has_api = "api" in str(vals[3]).lower() or "core_banking" in str(vals[3]).lower()
            if has_api:
                est_h = 8.0
            # Dead end → +1h de rediseño
            if vals[7]:
                est_h += 1.0
            vals.append(est_h)
            colors=[TEXT,TEXT,BLUE,DIM,DIM,
                    RED if vals[5]=="—" else TEXT,
                    TEXT,
                    RED if vals[7] else TEXT,
                    TEAL]
            row=data_row(ws,row,vals,col_start=2,alt=idx%2==0,colors=colors)
        # Total horas footer
        ws.row_dimensions[row].height=20
        ws.cell(row=row,column=2,value="TOTAL ESTIMATED MIGRATION HOURS")
        ws.cell(row=row,column=2).font=Font(name="Arial",bold=True,size=9,color=WHITE)
        ws.cell(row=row,column=2).fill=hf(SURFACE)
        ws.merge_cells(f"B{row}:H{row}")
        total_cell=ws.cell(row=row,column=10,value=f"=SUM(J6:J{row-1})")
        total_cell.font=Font(name="Arial",bold=True,size=11,color=TEAL)
        total_cell.fill=hf(SURFACE)
        total_cell.alignment=Alignment(horizontal="center",vertical="center")
        ws.cell(row=row,column=9,value="hrs")
        ws.cell(row=row,column=9).font=Font(name="Arial",size=8,color=DIM)
        ws.cell(row=row,column=9).fill=hf(SURFACE)
        ws.cell(row=row,column=9).alignment=Alignment(horizontal="right",vertical="center")
        row+=2
        # Disclaimer
        ws.merge_cells(f"B{row}:J{row}")
        disc = ws.cell(row=row,column=2,
            value="Estimaciones basadas en metodología de integrador certificado Genesys (Cyara/TTEC Digital/PS Genesys). "
                  "No representan datos oficiales de Genesys. El esfuerzo real depende del entorno técnico del cliente.")
        disc.font=Font(name="Arial",size=7,italic=True,color=DIM)
        disc.fill=hf(BG)
        disc.alignment=Alignment(wrap_text=True,vertical="center")
        ws.row_dimensions[row].height=22
        row+=1
    else:
        c=ws.cell(row=row,column=2,value="Node detail not available — run analysis to populate")
        c.font=Font(name="Arial",size=8,color=DIM,italic=True)
        c.fill=hf(SURFACE); row+=1

    # ── S2: EXTERNAL DEPENDENCIES ─────────────────────────────────────────────
    row=spacer(ws,row)
    row=section_header(ws,row,2,9,"§2 · EXTERNAL DEPENDENCIES",BLUE)
    row=table_header(ws,row,["Type","Service / Name","Count","Notes"])

    deps=[
        ("Data APIs",   ", ".join(inv.get("data_services",[])) or "None",  len(inv.get("data_services",[])), "Requires data reconect in Cloud"),
        ("Auth Services",", ".join(inv.get("auth_services",[])) or "None", len(inv.get("auth_services",[])), "Validate Genesys Cloud Auth compatibility"),
        ("Agent Queues", ", ".join(inv.get("unique_queues",[])) or "None",  len(inv.get("unique_queues",[])), "Queue routing config must be recreated"),
        ("TTS Variables",", ".join(f"{{{v}}}" for v in inv.get("dynamic_variables",[])) or "None", len(inv.get("dynamic_variables",[])), "Verify TTS runtime availability in Cloud"),
    ]
    for idx,(dtype,name,count,note) in enumerate(deps):
        vals=[dtype,name,count,note]
        colors=[TEAL,TEXT,YELLOW if count>0 else TEXT,DIM]
        row=data_row(ws,row,vals,col_start=2,alt=idx%2==0,colors=colors)

    # ── S3: RISK ANALYSIS ─────────────────────────────────────────────────────
    row=spacer(ws,row)
    row=section_header(ws,row,2,9,"§3 · RISK ANALYSIS",ORANGE)
    row=table_header(ws,row,["Risk Item","Value","Severity","Description"])

    dead_ends=inv.get("dead_ends",[]) or []
    fallbacks=inv.get("missing_fallbacks",[]) or []
    loops=inv.get("loop_nodes",[]) or []
    dtmf=inv.get("dtmf_input_nodes",[]) or []
    voice=inv.get("speech_nodes",[]) or []
    dual_input = bool(dtmf and voice)

    risks=[
        ("Dead End Nodes",   len(dead_ends),   "HIGH"   if dead_ends else "NONE",  ", ".join(dead_ends[:5]) or "None detected"),
        ("Missing Fallbacks",len(fallbacks),   "CRITICAL" if fallbacks else "NONE",", ".join(fallbacks[:5]) or "All nodes have fallbacks"),
        ("Loop Nodes",       len(loops),       "MEDIUM" if loops else "NONE",      ", ".join(loops[:5]) or "No loops detected"),
        ("Dual-Input (DTMF+Speech)",1 if dual_input else 0,"MEDIUM" if dual_input else "NONE","Requires dual recognition engine in Cloud"),
        ("Flow Depth (hops)",inv.get("flow_depth",0),"HIGH" if inv.get("flow_depth",0)>8 else "LOW","Max hops from entry to exit"),
        ("Entry Node Defined",1 if inv.get("entry_node_id") else 0,"CRITICAL" if not inv.get("entry_node_id") else "NONE","Required for Genesys Cloud migration"),
    ]
    sev_color={"CRITICAL":RED,"HIGH":ORANGE,"MEDIUM":YELLOW,"LOW":GREEN,"NONE":DIM}
    for idx,(item,val,sev,desc) in enumerate(risks):
        colors=[TEXT,YELLOW if val>0 else TEXT,sev_color.get(sev,TEXT),DIM]
        row=data_row(ws,row,[item,val,sev,desc],col_start=2,alt=idx%2==0,colors=colors)

    # ── S4: MIGRATION BREAKDOWN ───────────────────────────────────────────────
    row=spacer(ws,row)
    row=section_header(ws,row,2,9,"§4 · MIGRATION COMPLEXITY BREAKDOWN",PURPLE)
    row=table_header(ws,row,["Dimension","Score","Max","Pct","Key Drivers"])
    bd=inv.get("migration_score_breakdown",{})
    dim_info=[
        ("D1 · Graph Complexity","D1_grafo",25,BLUE),
        ("D2 · External Dependencies","D2_dependencias",25,TEAL),
        ("D3 · Business Risk","D3_riesgo",20,RED),
        ("D4 · Volume & Scale","D4_escala",15,YELLOW),
        ("D5 · Testing Effort","D5_testing",15,PURPLE),
    ]
    for idx,(label,key,max_v,color) in enumerate(dim_info):
        d=bd.get(key,{})
        sc=d.get("score",0); pct=f"{round(sc/max_v*100)}%" if max_v else "—"
        drivers=d.get("drivers","") or ""
        row=data_row(ws,row,[label,sc,max_v,pct,drivers],col_start=2,alt=idx%2==0,
                     colors=[color,color,DIM,YELLOW if sc>max_v*0.6 else TEXT,DIM])

    # Total
    total_ms=inv.get("migration_complexity_score",0)
    ws.row_dimensions[row].height=22
    ws.cell(row=row,column=2,value="TOTAL MIGRATION SCORE").font=Font(name="Arial",bold=True,size=9,color=WHITE)
    ws.cell(row=row,column=2).fill=hf(SURFACE)
    ws.cell(row=row,column=3,value=total_ms).font=Font(name="Arial",bold=True,size=12,color=mig_color(ml))
    ws.cell(row=row,column=3).fill=hf(SURFACE)
    ws.cell(row=row,column=3).alignment=Alignment(horizontal="center",vertical="center")
    ws.cell(row=row,column=4,value=100).font=Font(name="Arial",size=9,color=DIM)
    ws.cell(row=row,column=4).fill=hf(SURFACE)
    ws.cell(row=row,column=4).alignment=Alignment(horizontal="center",vertical="center")
    ws.cell(row=row,column=5,value=f"→ {ml}").font=Font(name="Arial",bold=True,size=10,color=mig_color(ml))
    ws.cell(row=row,column=5).fill=hf(SURFACE)
    row+=1

    # ── S5: FINDINGS & ACTION PLAN ────────────────────────────────────────────
    findings=r.get("critical_issues",[]) or []
    recommendations=r.get("recommendations",[]) or []
    if findings or recommendations:
        row=spacer(ws,row)
        row=section_header(ws,row,2,9,"§5 · FINDINGS & ACTION PLAN",RED)
        if findings:
            row=table_header(ws,row,["#","Finding","Severity"])
            for idx,f in enumerate(findings,1):
                row=data_row(ws,row,[idx,f,"CRITICAL"],col_start=2,alt=idx%2==0,
                             colors=[DIM,TEXT,RED])
        if recommendations:
            row=spacer(ws,row,6)
            row=table_header(ws,row,["#","Recommendation","Priority"])
            for idx,rec in enumerate(recommendations,1):
                row=data_row(ws,row,[idx,rec,"HIGH"],col_start=2,alt=idx%2==0,
                             colors=[DIM,TEXT,YELLOW])


# ── SHEET BUILDER: MIGRATION BREAKDOWN ────────────────────────────────────────
def build_migration_sheet(wb, data):
    ws=wb.create_sheet("Migration Breakdown")
    ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor=RED
    ws.freeze_panes="A5"; bg_all(ws)
    set_cols(ws,[3,28,10,14,18,20,16,14,14])
    ws.row_dimensions[2].height=34; ws.row_dimensions[4].height=20
    ws["B2"].value="Migration Complexity — 5 Dimensions"
    ws["B2"].font=Font(name="Arial",bold=True,size=13,color=WHITE)
    ws["B2"].alignment=Alignment(vertical="center")
    dim_keys=["D1_grafo","D2_dependencias","D3_riesgo","D4_escala","D5_testing"]
    dim_lbl=["Graph (25)","Ext.Deps (25)","Biz Risk (20)","Scale (15)","Testing (15)"]
    dim_clr=[BLUE,TEAL,RED,YELLOW,PURPLE]
    h3=["","Flow","TOTAL","Level"]+dim_lbl
    for col,h in enumerate(h3,1):
        c=ws.cell(row=4,column=col,value=h)
        c.font=Font(name="Arial",bold=True,size=8,color=dim_clr[col-5] if col>=5 else DIM)
        c.fill=hf(SURFACE); c.border=bb()
        c.alignment=Alignment(horizontal="left" if col<=2 else "center",vertical="center")
    for idx,r in enumerate(data):
        row=5+idx; ws.row_dimensions[row].height=20
        inv=r.get("inventory",{}); bd=inv.get("migration_score_breakdown",{})
        fname=r["filename"].replace(".yaml","").replace(".yml","")
        ml=inv.get("migration_level","—"); ms=inv.get("migration_complexity_score",0)
        rbg=CARD if idx%2==0 else SURFACE
        rv=["",fname,ms,ml]+[bd.get(k,{}).get("score",0) for k in dim_keys]
        for col,val in enumerate(rv,1):
            c=ws.cell(row=row,column=col,value=val)
            c.fill=hf(rbg); c.border=bb(BORDER)
            c.alignment=Alignment(horizontal="left" if col<=2 else "center",vertical="center")
            c.font=Font(name="Arial",size=9,color=TEXT if col<=4 else dim_clr[col-5])
        ws.cell(row=row,column=3).font=Font(name="Arial",bold=True,size=11,color=mig_color(ml))
        ws.cell(row=row,column=4).font=Font(name="Arial",size=8,bold=True,color=mig_color(ml))


# ── MAIN FUNCTION ─────────────────────────────────────────────────────────────
def generar_portfolio_excel_v2(results, raw_yamls=None):
    """
    results: lista de dicts con score, inventory, filename, summary, critical_issues, recommendations
    raw_yamls: dict {filename: yaml_dict} opcional para desglose de nodos
    """
    data = sorted([r for r in results if "error" not in r],
                  key=lambda x: x.get("score",0), reverse=True)
    today = date.today().strftime("%d/%m/%Y")
    raw_yamls = raw_yamls or {}
    wb = Workbook()

    build_overview(wb, data, today)
    build_migration_sheet(wb, data)
    for r in data:
        raw = raw_yamls.get(r["filename"])
        build_flow_sheet(wb, r, raw_yaml=raw)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.getvalue()





def ivr_loading_panel(current: int, total: int, current_name: str = "", phases: list = None) -> str:
    """
    Loading panel nivel Apple — grafo IVR con nodos tipados que se iluminan
    y conectan mientras avanza el análisis. Diseño oscuro, tipografía Syne.
    phases: lista de strings para los pasos (solo para modo individual)
    """
    pct = int((current / max(total, 1)) * 100) if total > 1 else min(current * 25, 95)

    # Grafo IVR — arquitectura realista de un flujo bancario tipo
    nodes = [
        # id, x%, y%, tipo, label, radio
        ("start",  50, 8,  "ENTRY",    "ENTRY",  8),
        ("auth",   28, 25, "AUTH",     "AUTH",   6),
        ("menu1",  72, 25, "MENU",     "MENU",   6),
        ("api1",   14, 45, "API",      "API",    5),
        ("task1",  38, 45, "TASK",     "TASK",   5),
        ("cond1",  62, 45, "COND",     "COND",   5),
        ("xfer1",  86, 45, "XFER",     "XFER",   5),
        ("vm1",    20, 65, "VOICE",    "TTS",    4),
        ("exit1",  44, 65, "EXIT",     "EXIT",   4),
        ("exit2",  68, 65, "EXIT",     "EXIT",   4),
        ("exit3",  88, 65, "EXIT",     "EXIT",   4),
    ]
    edges = [
        ("start","auth"), ("start","menu1"),
        ("auth","api1"), ("auth","task1"),
        ("menu1","cond1"), ("menu1","xfer1"),
        ("api1","vm1"), ("task1","exit1"),
        ("cond1","exit2"), ("xfer1","exit3"),
        ("vm1","exit1"),
    ]
    TYPE_COLOR = {
        "ENTRY": "#00D4AA", "AUTH": "#F85149", "MENU": "#00A8FF",
        "API":   "#F0883E", "TASK": "#A78BFA", "COND": "#D29922",
        "XFER":  "#00D4AA", "VOICE":"#3FB950", "EXIT": "#4B5568",
    }

    W, H = 360, 80
    lit = max(1, int(len(nodes) * pct / 100))

    # Edges SVG
    svg_e = ""
    for (a, b) in edges:
        n1 = next(n for n in nodes if n[0]==a)
        n2 = next(n for n in nodes if n[0]==b)
        x1,y1 = int(n1[1]*W/100), int(n1[2]*H/100)
        x2,y2 = int(n2[1]*W/100), int(n2[2]*H/100)
        i1 = next(i for i,n in enumerate(nodes) if n[0]==a)
        i2 = next(i for i,n in enumerate(nodes) if n[0]==b)
        lit_edge = i1 < lit and i2 < lit
        c = TYPE_COLOR.get(n2[3], "#4B5568")
        svg_e += '<line x1="%d" y1="%d" x2="%d" y2="%d" stroke="%s" stroke-width="%.1f" opacity="%.1f"/>' % (
            x1, y1, x2, y2, c if lit_edge else "#1C2030", 1.2 if lit_edge else 0.8, 0.8 if lit_edge else 1.0)

    # Nodes SVG
    svg_n = ""
    for i, (nid, nx, ny, ntype, nlabel, nr) in enumerate(nodes):
        cx, cy = int(nx*W/100), int(ny*H/100)
        c = TYPE_COLOR.get(ntype, "#4B5568")
        is_lit = i < lit
        is_cur = i == lit - 1
        op = "1" if is_lit else "0.12"
        if is_cur:
            # Outer pulse ring
            svg_n += '<circle cx="%d" cy="%d" r="%d" fill="none" stroke="%s" stroke-width="1.5" opacity="0.3"><animate attributeName="r" values="%d;%d;%d" dur="1s" repeatCount="indefinite"/><animate attributeName="opacity" values="0.3;0;0.3" dur="1s" repeatCount="indefinite"/></circle>' % (
                cx, cy, nr+5, c, nr+5, nr+10, nr+5)
            svg_n += '<circle cx="%d" cy="%d" r="%d" fill="%s" filter="drop-shadow(0 0 4px %s)"/>' % (cx, cy, nr, c, c)
        else:
            svg_n += '<circle cx="%d" cy="%d" r="%d" fill="%s" opacity="%s"/>' % (cx, cy, nr, c, op)
        svg_n += '<text x="%d" y="%d" text-anchor="middle" fill="%s" opacity="%s" font-family="DM Mono,monospace" font-size="5.5" font-weight="600">%s</text>' % (
            cx, cy+nr+8, c, op, nlabel)

    # Nombre del archivo actual (recortado)
    fname_short = current_name.replace(".yaml","").replace(".yml","")[:34]

    # Render de fases si es modo individual
    phases_html = ""
    if phases:
        for idx2, ph in enumerate(phases):
            if idx2 < current:
                style = "color:#00D4AA;font-size:0.68rem;"
                icon = "✓"
            elif idx2 == current:
                style = "color:#00A8FF;font-size:0.68rem;"
                icon = "▶"
            else:
                style = "color:#252D3D;font-size:0.68rem;"
                icon = "○"
            phases_html += '<div style="font-family:DM Mono,monospace;%s padding:3px 0;border-bottom:1px solid #0D1017;">%s %s</div>' % (style, icon, ph)

    counter_html = '<span style="font-family:Syne,sans-serif;font-size:1.1rem;font-weight:800;color:#00D4AA;">%d/%d</span>' % (current, total) if total > 1 else '<span style="font-family:DM Mono,monospace;font-size:0.65rem;color:#4B5568;">ANALYZING...</span>'

    return (
        '<div style="background:#0B0D14;border:1px solid #1C2030;border-radius:12px;padding:1rem 1.25rem;">'
        '<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:0.5rem;">'
        '<span style="font-family:DM Mono,monospace;font-size:0.6rem;color:#4B5568;letter-spacing:0.1em;">PROCESSING FLOW</span>'
        + counter_html
        + '</div>'
        '<svg viewBox="0 0 360 100" xmlns="http://www.w3.org/2000/svg" style="width:100%;height:80px;display:block;">'
        + svg_e + svg_n
        + '</svg>'
        + (phases_html if phases_html else '')
        + '<div style="margin-top:0.6rem;">'
          '<div style="display:flex;justify-content:space-between;font-family:DM Mono,monospace;font-size:0.6rem;color:#4B5568;margin-bottom:3px;">'
          '<span style="color:#E8EDF5;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:78%;">' + fname_short + '</span>'
          '<span style="color:#00D4AA;">' + str(pct) + '%</span>'
          '</div>'
          '<div style="background:#0E1118;border-radius:2px;height:3px;">'
          '<div style="background:linear-gradient(90deg,#00D4AA,#00A8FF);height:100%;width:' + str(pct) + '%;border-radius:2px;transition:width 0.3s ease;"></div>'
          '</div></div></div>'
    )


def generar_portfolio_pdf(results, flows_map) -> bytes:
    """PDF consolidado: portada portfolio + 1 página por flujo."""
    import io, json, re, os
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable, PageBreak)
    from reportlab.lib.units import cm
    from datetime import datetime

    # ── PALETA ────────────────────────────────────────────────────────────────
    navy    = HexColor('#1E3A5F')
    teal    = HexColor('#00D4AA')
    blue    = HexColor('#00A8FF')
    accent  = HexColor('#00D4AA')
    red     = HexColor('#F85149')
    orange  = HexColor('#F0883E')
    green   = HexColor('#3FB950')
    yellow  = HexColor('#D29922')
    white   = HexColor('#FFFFFF')
    light   = HexColor('#F8FAFC')
    border  = HexColor('#E2E8F0')
    dim     = HexColor('#64748B')
    dark    = HexColor('#1E293B')

    def score_color(s):
        return green if s >= 70 else (orange if s >= 40 else red)
    def mig_color(l):
        return {"SIMPLE":green,"MODERADO":yellow,"COMPLEJO":orange,"MUY COMPLEJO":red}.get(l, dim)
    def mig_bg(l):
        return {"SIMPLE":HexColor('#D1FAE5'),"MODERADO":HexColor('#FEF3C7'),
                "COMPLEJO":HexColor('#FFEDD5'),"MUY COMPLEJO":HexColor('#FEE2E2')}.get(l, light)

    # ── ESTILOS ───────────────────────────────────────────────────────────────
    brand  = ParagraphStyle('Brand', fontSize=9,  textColor=teal, fontName='Helvetica-Bold', spaceAfter=3)
    h1     = ParagraphStyle('H1',    fontSize=22, textColor=navy, fontName='Helvetica-Bold', spaceAfter=6)
    h2     = ParagraphStyle('H2',    fontSize=14, textColor=navy, fontName='Helvetica-Bold', spaceAfter=8,  spaceBefore=18)
    h3     = ParagraphStyle('H3',    fontSize=11, textColor=navy, fontName='Helvetica-Bold', spaceAfter=5,  spaceBefore=10)
    sub    = ParagraphStyle('Sub',   fontSize=10, textColor=dim,  spaceAfter=14)
    body   = ParagraphStyle('Body',  fontSize=9,  textColor=dark, spaceAfter=5,  leading=15)
    footer = ParagraphStyle('Foot',  fontSize=7,  textColor=dim,  alignment=1)
    small  = ParagraphStyle('Small', fontSize=8,  textColor=dim,  spaceAfter=3)

    ok = sorted([r for r in results if 'error' not in r],
                key=lambda x: x.get('score', 0), reverse=True)
    today = datetime.now().strftime('%d/%m/%Y')
    now   = datetime.now().strftime('%d/%m/%Y %H:%M')

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    story = []

    # ══════════════════════════════════════════════════════════════════════════
    # PÁGINA 1 — PORTADA DEL PORTFOLIO
    # ══════════════════════════════════════════════════════════════════════════
    story.append(Paragraph('OrchestrIA · IVR·IA', brand))
    story.append(Paragraph('Portfolio Analysis Report', h1))
    story.append(Paragraph(f'{len(ok)} flows analyzed  ·  {today}', sub))
    story.append(HRFlowable(width='100%', thickness=1, color=border, spaceAfter=14))

    # Tabla resumen portfolio
    avg_score = round(sum(r.get('score',0) for r in ok) / max(len(ok),1))
    levels = [r.get('inventory',{}).get('migration_level','—') for r in ok]
    level_counts = {l: levels.count(l) for l in set(levels)}

    summary_data = [
        ['Métrica', 'Valor'],
        ['Flows analizados', str(len(ok))],
        ['Score promedio', f'{avg_score}/100'],
        ['Flows con score ≥ 70 (GOOD)', str(sum(1 for r in ok if r.get('score',0) >= 70))],
        ['Flows con score < 40 (POOR)', str(sum(1 for r in ok if r.get('score',0) < 40))],
        ['Distribución de complejidad', '  ·  '.join(f'{v} {k}' for k,v in level_counts.items())],
    ]
    t = Table(summary_data, colWidths=[9*cm, 6*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), navy),
        ('TEXTCOLOR',  (0,0), (-1,0), white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTNAME',   (0,1), (0,-1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 9),
        ('PADDING',    (0,0), (-1,-1), 8),
        ('GRID',       (0,0), (-1,-1), 0.5, border),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [light, white]),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    # Ranking de flows
    story.append(Paragraph('Ranking de Flujos por Score de Calidad', h2))
    rank_data = [['#', 'Flujo', 'Score', 'Calidad', 'Nodos', 'Ext.Deps', 'Migración', 'Mig.Score']]
    for i, r in enumerate(ok, 1):
        inv = r.get('inventory', {})
        s   = r.get('score', 0)
        ml  = inv.get('migration_level', '—')
        ql  = 'GOOD' if s >= 70 else ('FAIR' if s >= 40 else 'POOR')
        rank_data.append([
            str(i),
            r['filename'].replace('.yaml','').replace('.yml','')[:30],
            f'{s}/100', ql,
            str(inv.get('total_nodes', 0)),
            str(inv.get('total_external_deps', 0)),
            ml,
            str(inv.get('migration_complexity_score', 0)),
        ])
    rt = Table(rank_data, colWidths=[0.6*cm,5.5*cm,1.5*cm,1.5*cm,1.2*cm,1.5*cm,2.5*cm,1.8*cm])
    tstyle = [
        ('BACKGROUND', (0,0), (-1,0), navy),
        ('TEXTCOLOR',  (0,0), (-1,0), white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('PADDING',    (0,0), (-1,-1), 5),
        ('GRID',       (0,0), (-1,-1), 0.3, border),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [light, white]),
        ('ALIGN',      (2,0), (-1,-1), 'CENTER'),
    ]
    # Color semafórico en columna Score
    for i, r in enumerate(ok, 1):
        s  = r.get('score', 0)
        ml = r.get('inventory',{}).get('migration_level','—')
        tstyle.append(('TEXTCOLOR', (2,i), (2,i), score_color(s)))
        tstyle.append(('FONTNAME',  (2,i), (2,i), 'Helvetica-Bold'))
        tstyle.append(('TEXTCOLOR', (6,i), (6,i), mig_color(ml)))
    rt.setStyle(TableStyle(tstyle))
    story.append(rt)

    # ══════════════════════════════════════════════════════════════════════════
    # PÁGINAS INDIVIDUALES — UNA POR FLUJO
    # ══════════════════════════════════════════════════════════════════════════
    for r in ok:
        story.append(PageBreak())
        inv   = r.get('inventory', {})
        score = r.get('score', 0)
        ml    = inv.get('migration_level', '—')
        ms    = inv.get('migration_complexity_score', 0)
        fname = r['filename'].replace('.yaml','').replace('.yml','')

        # Cabecera del flujo
        story.append(Paragraph('OrchestrIA · IVR·IA', brand))
        story.append(Paragraph(f'Informe de Auditoría de Flujo IVR', h1))
        story.append(Paragraph(f'{fname}  ·  {today}', sub))
        story.append(HRFlowable(width='100%', thickness=1, color=border, spaceAfter=12))

        # Tabla métricas
        metrics = [
            ['Métrica', 'Valor', 'Métrica', 'Valor'],
            ['Score de Calidad', f'{score}/100', 'Total Nodos', str(inv.get('total_nodes',0))],
            ['Self-Service Ratio', f"{inv.get('self_service_ratio',0)}%",
             'Transfers a Agente', str(inv.get('agent_transfers',0))],
            ['Deps. Externas', str(inv.get('total_external_deps',0)),
             'Complejidad Migración', ml],
            ['APIs de Datos', str(len(inv.get('data_services',[]))),
             'Servicios de Auth', str(len(inv.get('auth_services',[])))],
        ]
        mt = Table(metrics, colWidths=[4.5*cm, 3*cm, 4.5*cm, 3*cm])
        mt.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), navy),
            ('TEXTCOLOR',  (0,0), (-1,0), white),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTNAME',   (0,1), (0,-1), 'Helvetica-Bold'),
            ('FONTNAME',   (2,1), (2,-1), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 9),
            ('PADDING',    (0,0), (-1,-1), 7),
            ('GRID',       (0,0), (-1,-1), 0.4, border),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [light, white]),
            ('TEXTCOLOR',  (1,1), (1,1), score_color(score)),
            ('FONTNAME',   (1,1), (1,1), 'Helvetica-Bold'),
            ('TEXTCOLOR',  (3,3), (3,3), mig_color(ml)),
            ('FONTNAME',   (3,3), (3,3), 'Helvetica-Bold'),
        ]))
        story.append(mt)

        # Inventario
        story.append(Paragraph('Inventario del Flujo', h2))
        inv_rows = [['Componente', 'Detalle']]
        if inv.get('data_services'):
            inv_rows.append(['APIs de Datos', ', '.join(inv['data_services'])])
        if inv.get('auth_services'):
            inv_rows.append(['Servicios de Auth', ', '.join(inv['auth_services'])])
        if inv.get('unique_queues'):
            inv_rows.append(['Colas de Destino', ', '.join(inv['unique_queues'])])
        if inv.get('dynamic_variables'):
            inv_rows.append(['Variables TTS', ', '.join('{'+v+'}' for v in inv['dynamic_variables'])])
        inv_rows.append(['Menú / Transfer / Logic',
                         f"{inv.get('menu_nodes',0)} / {inv.get('transfer_nodes',0)} / {inv.get('task_nodes',0)}"])
        it = Table(inv_rows, colWidths=[5*cm, 10*cm])
        it.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), navy),
            ('TEXTCOLOR',  (0,0), (-1,0), white),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTNAME',   (0,1), (0,-1), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 8),
            ('PADDING',    (0,0), (-1,-1), 6),
            ('GRID',       (0,0), (-1,-1), 0.4, border),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [light, white]),
        ]))
        story.append(it)

        # Resumen ejecutivo
        summary = r.get('summary') or r.get('executive_summary') or ''
        if summary:
            story.append(Paragraph('Resumen Ejecutivo', h2))
            story.append(Paragraph(summary, body))

        # Hallazgos
        findings = r.get('critical_issues') or []
        if findings:
            story.append(Paragraph('Hallazgos Principales', h2))
            for i, f in enumerate(findings, 1):
                story.append(Paragraph(f'{i}. {f}', body))

        # Recomendaciones
        recs = r.get('improvements') or r.get('recommendations') or []
        if recs:
            story.append(Paragraph('Plan de Acción', h2))
            for i, rec in enumerate(recs, 1):
                story.append(Paragraph(f'Paso {i}: {rec}', body))

        # Migration Assessment
        story.append(Paragraph('Migration Assessment · Genesys Cloud', h2))
        story.append(Paragraph(f'Nivel de complejidad: {ml} ({ms}/100)', h3))
        flags = inv.get('migration_risk_flags') or []
        if flags:
            story.append(Paragraph('Riesgos identificados:', small))
            for flag in flags:
                story.append(Paragraph(f'• {flag}', body))

        # Footer
        story.append(Spacer(1, 0.5*cm))
        story.append(HRFlowable(width='100%', thickness=0.5, color=border, spaceAfter=6))
        story.append(Paragraph(
            f'Informe generado por OrchestrIA IVR·IA · {now} · Confidencial',
            footer))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()



def calcular_horas_estimadas(inv: dict) -> dict:
    """
    Estima horas de migración desde el inventory dict.
    Basado en metodología Cyara/TTEC Digital/PS Genesys.
    Devuelve dict con breakdown por categoría y total.
    """
    # Horas base por tipo de nodo (conteos del inventory)
    menu_h    = inv.get('menu_nodes',      0) * 3.0
    xfer_h    = inv.get('transfer_nodes',  0) * 4.0
    logic_h   = inv.get('task_nodes',      0) * 3.0
    voice_h   = len(inv.get('voicemail_nodes', [])) * 3.0

    # APIs/integraciones — 8h por data service, 4h por auth
    api_h     = len(inv.get('data_services',  [])) * 8.0
    auth_h    = len(inv.get('auth_services',  [])) * 4.0

    # Testing — variable dinámicas TTS requieren validación en Cloud runtime
    tts_h     = len(inv.get('dynamic_variables', [])) * 1.0

    # Riesgo: dead ends y fallbacks faltantes requieren rediseño
    risk_h    = (len(inv.get('dead_ends',         [])) +
                 len(inv.get('missing_fallbacks',  []))) * 1.0

    # Entry node faltante: +2h arquitectura de entrada
    entry_h   = 2.0 if not inv.get('entry_node_id') else 0.0

    # Categorías agrupadas
    routing_h     = menu_h + xfer_h + voice_h
    integration_h = api_h + auth_h + tts_h
    testing_h     = max(logic_h + risk_h + entry_h, 2.0)  # min 2h de testing

    total = routing_h + integration_h + testing_h
    days  = round(total / 8, 1)  # 8h/día de trabajo

    return {
        'routing_hours':     round(routing_h, 1),
        'integration_hours': round(integration_h, 1),
        'testing_hours':     round(testing_h, 1),
        'total_hours':       round(total, 1),
        'days_estimate':     days,
    }


def migration_hours_card(inv: dict, compact: bool = False) -> str:
    """
    Card HTML con estimación de horas de migración.
    compact=True para uso en batch expanders.
    """
    h = calcular_horas_estimadas(inv)
    total  = h['total_hours']
    days   = h['days_estimate']
    ml     = inv.get('migration_level', 'SIMPLE')

    ml_color = {
        'SIMPLE':      '#00D4AA',
        'MODERADO':    '#D29922',
        'COMPLEJO':    '#F0883E',
        'MUY COMPLEJO':'#F85149',
    }.get(ml, '#4B5568')

    if compact:
        return (
            f'''<div style="display:flex;align-items:center;gap:1rem;
                 background:#0B0D14;border:1px solid #1C2030;border-radius:8px;
                 padding:0.6rem 1rem;margin-top:0.5rem;">
              <div>
                <div style="font-family:DM Mono,monospace;font-size:0.55rem;
                     color:#4B5568;letter-spacing:0.08em;">EST. MIGRATION</div>
                <div style="font-family:Syne,sans-serif;font-weight:800;
                     font-size:1.4rem;color:{ml_color};line-height:1;">
                  {total}h
                </div>
                <div style="font-family:DM Mono,monospace;font-size:0.6rem;
                     color:#4B5568;">~{days} days</div>
              </div>
              <div style="flex:1;display:flex;flex-direction:column;gap:4px;">
                <div style="display:flex;justify-content:space-between;
                     font-family:DM Mono,monospace;font-size:0.6rem;">
                  <span style="color:#4B5568;">Routing</span>
                  <span style="color:#00A8FF;">{h["routing_hours"]}h</span>
                </div>
                <div style="display:flex;justify-content:space-between;
                     font-family:DM Mono,monospace;font-size:0.6rem;">
                  <span style="color:#4B5568;">Integration</span>
                  <span style="color:#F0883E;">{h["integration_hours"]}h</span>
                </div>
                <div style="display:flex;justify-content:space-between;
                     font-family:DM Mono,monospace;font-size:0.6rem;">
                  <span style="color:#4B5568;">Testing & Risk</span>
                  <span style="color:#A78BFA;">{h["testing_hours"]}h</span>
                </div>
              </div>
            </div>'''
        )

    bar_total = max(total, 1)
    r_pct = round(h["routing_hours"]     / bar_total * 100)
    i_pct = round(h["integration_hours"] / bar_total * 100)
    t_pct = 100 - r_pct - i_pct

    disclaimer = (
        '<div style="font-family:DM Mono,monospace;font-size:0.55rem;color:#2A3545;'        'margin-top:0.75rem;line-height:1.5;">'        'Estimación basada en metodología Cyara · TTEC Digital · PS Genesys. '        'No representa datos oficiales de Genesys Systems Inc.</div>'
    )

    return f'''
<div style="background:#0B0D14;border:1px solid #1C2030;border-radius:10px;
     padding:1.25rem 1.5rem;margin-top:0.5rem;">
  <div style="display:flex;justify-content:space-between;align-items:flex-start;
       margin-bottom:1rem;">
    <div>
      <div style="font-family:DM Mono,monospace;font-size:0.6rem;
           color:#4B5568;letter-spacing:0.08em;margin-bottom:4px;">
        MIGRATION EFFORT ESTIMATE · ENGAGE → CLOUD
      </div>
      <div style="display:flex;align-items:baseline;gap:0.5rem;">
        <span style="font-family:Syne,sans-serif;font-weight:800;
              font-size:2.2rem;color:{ml_color};line-height:1;">{total}h</span>
        <span style="font-family:DM Mono,monospace;font-size:0.75rem;
              color:#4B5568;">~{days} working days</span>
      </div>
    </div>
    <div style="text-align:right;">
      <div style="font-family:DM Mono,monospace;font-size:0.55rem;
           color:#4B5568;margin-bottom:4px;">COMPLEXITY</div>
      <div style="font-family:Syne,sans-serif;font-weight:800;
           font-size:0.85rem;color:{ml_color};">{ml}</div>
    </div>
  </div>

  <div style="display:flex;height:6px;border-radius:3px;overflow:hidden;
       margin-bottom:0.75rem;gap:2px;">
    <div style="width:{r_pct}%;background:#00A8FF;border-radius:3px 0 0 3px;"></div>
    <div style="width:{i_pct}%;background:#F0883E;"></div>
    <div style="width:{t_pct}%;background:#A78BFA;border-radius:0 3px 3px 0;"></div>
  </div>

  <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:0.5rem;">
    <div style="background:#0E1118;border-radius:6px;padding:0.5rem 0.75rem;">
      <div style="font-family:DM Mono,monospace;font-size:0.55rem;
           color:#4B5568;margin-bottom:2px;">ROUTING</div>
      <div style="font-family:Syne,sans-serif;font-weight:700;
           font-size:1.1rem;color:#00A8FF;">{h["routing_hours"]}h</div>
      <div style="font-family:DM Mono,monospace;font-size:0.55rem;color:#4B5568;">
        menus · transfers · voicemail
      </div>
    </div>
    <div style="background:#0E1118;border-radius:6px;padding:0.5rem 0.75rem;">
      <div style="font-family:DM Mono,monospace;font-size:0.55rem;
           color:#4B5568;margin-bottom:2px;">INTEGRATION</div>
      <div style="font-family:Syne,sans-serif;font-weight:700;
           font-size:1.1rem;color:#F0883E;">{h["integration_hours"]}h</div>
      <div style="font-family:DM Mono,monospace;font-size:0.55rem;color:#4B5568;">
        APIs · auth · TTS vars
      </div>
    </div>
    <div style="background:#0E1118;border-radius:6px;padding:0.5rem 0.75rem;">
      <div style="font-family:DM Mono,monospace;font-size:0.55rem;
           color:#4B5568;margin-bottom:2px;">TESTING & RISK</div>
      <div style="font-family:Syne,sans-serif;font-weight:700;
           font-size:1.1rem;color:#A78BFA;">{h["testing_hours"]}h</div>
      <div style="font-family:DM Mono,monospace;font-size:0.55rem;color:#4B5568;">
        logic · dead ends · fixes
      </div>
    </div>
  </div>
  {disclaimer}
</div>
'''


def portfolio_summary_card(results: list) -> str:
    """Card de resumen del portfolio con total de horas de migración."""
    ok = [r for r in results if 'error' not in r]
    if not ok:
        return ''

    total_h = 0.0
    total_days = 0.0
    scores = []
    levels = []
    for r in ok:
        inv = r.get('inventory', {})
        h = calcular_horas_estimadas(inv)
        total_h += h['total_hours']
        scores.append(r.get('score', 0))
        levels.append(inv.get('migration_level', 'SIMPLE'))

    total_days = round(total_h / 8, 1)
    avg_score  = round(sum(scores) / len(scores)) if scores else 0
    level_counts = {l: levels.count(l) for l in ['SIMPLE','MODERADO','COMPLEJO','MUY COMPLEJO']}
    sc = '#00D4AA' if avg_score >= 70 else ('#D29922' if avg_score >= 40 else '#F85149')

    level_pills = ''
    lc_colors = {'SIMPLE':'#00D4AA','MODERADO':'#D29922','COMPLEJO':'#F0883E','MUY COMPLEJO':'#F85149'}
    for lv, cnt in level_counts.items():
        if cnt:
            c = lc_colors.get(lv,'#4B5568')
            level_pills += (
                f'<span style="font-family:DM Mono,monospace;font-size:0.6rem;'                f'color:{c};background:{c}18;border:1px solid {c}40;'                f'border-radius:4px;padding:2px 8px;">{cnt} {lv}</span> '
            )

    return f'''
<div style="background:linear-gradient(135deg,#0B0D14 0%,#0E1118 100%);
     border:1px solid #1C2030;border-radius:12px;padding:1.5rem;
     margin-bottom:1.5rem;position:relative;overflow:hidden;">
  <div style="position:absolute;top:0;right:0;width:200px;height:100%;
       background:linear-gradient(90deg,transparent,#00D4AA08);
       pointer-events:none;"></div>

  <div style="font-family:DM Mono,monospace;font-size:0.6rem;
       color:#4B5568;letter-spacing:0.1em;margin-bottom:0.75rem;">
    PORTFOLIO MIGRATION SUMMARY · {len(ok)} FLOWS ANALYZED
  </div>

  <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:1rem;
       margin-bottom:1rem;">
    <div>
      <div style="font-family:DM Mono,monospace;font-size:0.55rem;
           color:#4B5568;margin-bottom:2px;">TOTAL EFFORT</div>
      <div style="font-family:Syne,sans-serif;font-weight:800;
           font-size:2rem;color:#00D4AA;line-height:1;">{round(total_h,0):.0f}h</div>
      <div style="font-family:DM Mono,monospace;font-size:0.6rem;
           color:#4B5568;">~{total_days} working days</div>
    </div>
    <div>
      <div style="font-family:DM Mono,monospace;font-size:0.55rem;
           color:#4B5568;margin-bottom:2px;">AVG QUALITY SCORE</div>
      <div style="font-family:Syne,sans-serif;font-weight:800;
           font-size:2rem;color:{sc};line-height:1;">{avg_score}</div>
      <div style="font-family:DM Mono,monospace;font-size:0.6rem;
           color:#4B5568;">out of 100</div>
    </div>
    <div>
      <div style="font-family:DM Mono,monospace;font-size:0.55rem;
           color:#4B5568;margin-bottom:2px;">FLOWS ANALYZED</div>
      <div style="font-family:Syne,sans-serif;font-weight:800;
           font-size:2rem;color:#E8EDF5;line-height:1;">{len(ok)}</div>
      <div style="font-family:DM Mono,monospace;font-size:0.6rem;
           color:#4B5568;">ready for export</div>
    </div>
    <div>
      <div style="font-family:DM Mono,monospace;font-size:0.55rem;
           color:#4B5568;margin-bottom:2px;">AVG PER FLOW</div>
      <div style="font-family:Syne,sans-serif;font-weight:800;
           font-size:2rem;color:#00A8FF;line-height:1;">{round(total_h/len(ok),0):.0f}h</div>
      <div style="font-family:DM Mono,monospace;font-size:0.6rem;
           color:#4B5568;">per flow average</div>
    </div>
  </div>

  <div style="display:flex;flex-wrap:wrap;gap:6px;">
    {level_pills}
  </div>
</div>
'''



# ── BENCHMARKS SECTORIALES ────────────────────────────────────────────────────
# Basado en Contact Babel Industry Reports 2023-2024 + Genesys State of CX 2024
# y experiencia de campo de integradores certificados
SECTOR_BENCHMARKS = {
    "banking": {
        "label": "Banca",
        "self_service_avg": 45,
        "self_service_top": 65,
        "quality_avg": 62,
        "quality_top": 78,
        "migration_avg": 38,
        "note": "Alta regulación (PCI-DSS, GDPR). Auth obligatoria en operaciones.",
    },
    "insurance": {
        "label": "Seguros",
        "self_service_avg": 38,
        "self_service_top": 58,
        "quality_avg": 58,
        "quality_top": 74,
        "migration_avg": 32,
        "note": "Flujos de siniestros complejos. Grabación obligatoria.",
    },
    "telco": {
        "label": "Telecomunicaciones",
        "self_service_avg": 55,
        "self_service_top": 72,
        "quality_avg": 64,
        "quality_top": 80,
        "migration_avg": 28,
        "note": "Mayor madurez en autoservicio. Pocas integraciones críticas.",
    },
    "public": {
        "label": "Administración Pública",
        "self_service_avg": 28,
        "self_service_top": 45,
        "quality_avg": 52,
        "quality_top": 68,
        "migration_avg": 22,
        "note": "Flujos simples pero volúmenes altos y horarios restrictivos.",
    },
    "ecommerce": {
        "label": "eCommerce / Retail",
        "self_service_avg": 48,
        "self_service_top": 68,
        "quality_avg": 60,
        "quality_top": 76,
        "migration_avg": 25,
        "note": "Alta estacionalidad. APIs de estado de pedido frecuentes.",
    },
    "healthcare": {
        "label": "Sanidad",
        "self_service_avg": 32,
        "self_service_top": 50,
        "quality_avg": 55,
        "quality_top": 70,
        "migration_avg": 35,
        "note": "HIPAA/LOPD. Grabación regulada. Transfers críticos.",
    },
}


def score_explanation(analysis: dict) -> str:
    """
    Genera explicación textual determinista del score y la complejidad
    de migración. Sin LLM — puramente desde el inventory.
    """
    inv   = analysis.get("inventory", {})
    score = analysis.get("score", 0)
    bd    = inv.get("migration_score_breakdown", {})
    ml    = inv.get("migration_level", "SIMPLE")
    ms    = inv.get("migration_complexity_score", 0)

    # ── Drivers del Quality Score ─────────────────────────────────────────────
    score_drivers = []
    menus     = inv.get("menu_nodes", 0)
    menus_bad = inv.get("menus_without_handlers", [])
    fallbacks = inv.get("missing_fallbacks", [])
    dead      = inv.get("dead_ends", [])
    self_svc  = inv.get("self_service_ratio", 0)
    apis      = inv.get("data_services", [])
    auth      = inv.get("auth_services", [])
    dvars     = inv.get("dynamic_variables", [])

    if isinstance(menus_bad, list) and len(menus_bad) > 0:
        score_drivers.append(
            f"{len(menus_bad)}/{menus} menús sin timeout ni handler de no-input"
        )
    if isinstance(fallbacks, list) and len(fallbacks) > 0:
        score_drivers.append(
            f"{len(fallbacks)} transfer(s) sin fallback — desconexión silenciosa posible"
        )
    if isinstance(dead, list) and len(dead) > 0:
        score_drivers.append(
            f"{len(dead)} dead end(s) — nodos sin salida definida"
        )
    if self_svc < 30:
        score_drivers.append(
            f"Autoservicio {self_svc}% — muy por debajo del benchmark sectorial (40-55%)"
        )
    if not inv.get("entry_node_id"):
        score_drivers.append("Entry node no definido — punto de entrada ambiguo")

    # ── Drivers del Migration Score ───────────────────────────────────────────
    mig_drivers = []
    d1 = bd.get("D1_grafo",        {}).get("score", 0)
    d2 = bd.get("D2_dependencias", {}).get("score", 0)
    d3 = bd.get("D3_riesgo",       {}).get("score", 0)
    d4 = bd.get("D4_escala",       {}).get("score", 0)
    d5 = bd.get("D5_testing",      {}).get("score", 0)

    # Identificar las dimensiones más pesadas
    dims = [
        (d1, "complejidad de grafo",     25),
        (d2, "dependencias externas",    25),
        (d3, "riesgo de negocio",        20),
        (d4, "volumen y escala",         15),
        (d5, "esfuerzo de testing",      15),
    ]
    # Dimensiones que superan el 60% de su máximo
    high_dims = [(score, label) for score, label, max_v in dims
                 if max_v > 0 and score / max_v >= 0.6]

    if apis:
        mig_drivers.append(
            f"{len(apis)} integración(es) de datos ({', '.join(apis[:2])})"
        )
    if auth:
        mig_drivers.append(
            f"{len(auth)} auth service(s) — requieren validación OAuth/SAML Cloud"
        )
    if dvars:
        mig_drivers.append(
            f"{len(dvars)} variable(s) TTS dinámica(s) — verificar runtime Cloud"
        )
    if isinstance(dead, list) and dead:
        mig_drivers.append(f"{len(dead)} dead end(s) — rediseño previo a migración")
    if not inv.get("entry_node_id"):
        mig_drivers.append("Entry node no definido — prerequisito arquitectónico")

    # ── Render HTML ───────────────────────────────────────────────────────────
    score_color_map = {
        True:  "#F85149",   # score < 40
        False: "#D29922",   # score < 70
    }
    sc = "#F85149" if score < 40 else ("#D29922" if score < 70 else "#00D4AA")
    mc = {"SIMPLE":"#00D4AA","MODERADO":"#D29922",
          "COMPLEJO":"#F0883E","MUY COMPLEJO":"#F85149"}.get(ml,"#4B5568")

    def driver_pill(text, color="#4B5568"):
        return (
            f'<div style="display:flex;align-items:flex-start;gap:8px;'            f'padding:5px 0;border-bottom:1px solid #0E1118;">'            f'<span style="color:{color};font-size:0.7rem;margin-top:1px;">▸</span>'            f'<span style="font-family:Plus Jakarta Sans,sans-serif;'            f'font-size:0.75rem;color:#7A8BA5;line-height:1.4;">{text}</span>'            f'</div>'
        )

    score_section = ""
    if score_drivers:
        pills = "".join(driver_pill(d, sc) for d in score_drivers)
        score_section = (
            f'<div style="margin-bottom:0.75rem;">'            f'<div style="font-family:DM Mono,monospace;font-size:0.6rem;'            f'color:#4B5568;letter-spacing:0.08em;margin-bottom:6px;">'            f'QUALITY SCORE — FACTORES PRINCIPALES</div>'            f'{pills}</div>'
        )

    mig_section = ""
    if mig_drivers:
        pills = "".join(driver_pill(d, mc) for d in mig_drivers)
        mig_section = (
            f'<div>'            f'<div style="font-family:DM Mono,monospace;font-size:0.6rem;'            f'color:#4B5568;letter-spacing:0.08em;margin-bottom:6px;">'            f'MIGRACIÓN {ml} — DRIVERS</div>'            f'{pills}</div>'
        )

    if not score_section and not mig_section:
        return ""

    return (
        f'<div style="background:#0B0D14;border:1px solid #1C2030;'        f'border-radius:10px;padding:1rem 1.25rem;margin-top:0.5rem;">'        f'{score_section}{mig_section}'        f'</div>'
    )


def benchmark_card(inv: dict, sector: str = "banking") -> str:
    """
    Card de comparación contra benchmark sectorial.
    sector: banking | insurance | telco | public | ecommerce | healthcare
    """
    b = SECTOR_BENCHMARKS.get(sector, SECTOR_BENCHMARKS["banking"])
    self_svc = inv.get("self_service_ratio", 0)
    ms       = inv.get("migration_complexity_score", 0)

    def gap_color(val, avg, top):
        if val >= top:   return "#00D4AA"
        if val >= avg:   return "#D29922"
        return "#F85149"

    def pct_bar(val, max_val=100, color="#00D4AA"):
        w = min(int(val / max_val * 100), 100)
        return (
            f'<div style="flex:1;background:#0E1118;border-radius:2px;height:4px;">'            f'<div style="width:{w}%;background:{color};height:100%;border-radius:2px;"></div>'            f'</div>'
        )

    ss_color  = gap_color(self_svc, b["self_service_avg"], b["self_service_top"])
    mig_color_b = "#00D4AA" if ms <= b["migration_avg"] else (
                  "#D29922" if ms <= b["migration_avg"]*1.5 else "#F85149")

    return f'''
<div style="background:#0B0D14;border:1px solid #1C2030;border-radius:10px;
     padding:1rem 1.25rem;margin-top:0.5rem;">
  <div style="display:flex;justify-content:space-between;align-items:center;
       margin-bottom:0.75rem;">
    <div style="font-family:DM Mono,monospace;font-size:0.6rem;
         color:#4B5568;letter-spacing:0.08em;">
      BENCHMARK SECTORIAL · {b["label"].upper()}
    </div>
    <div style="font-family:DM Mono,monospace;font-size:0.6rem;color:#3D4D66;">
      Fuente: Contact Babel 2024 · Genesys State of CX 2024
    </div>
  </div>

  <div style="display:grid;grid-template-columns:1fr 1fr;gap:0.75rem;margin-bottom:0.75rem;">
    <div style="background:#0E1118;border-radius:8px;padding:0.6rem 0.8rem;">
      <div style="font-family:DM Mono,monospace;font-size:0.55rem;color:#4B5568;margin-bottom:4px;">
        SELF-SERVICE RATIO
      </div>
      <div style="display:flex;align-items:baseline;gap:0.4rem;margin-bottom:6px;">
        <span style="font-family:Syne,sans-serif;font-weight:800;
              font-size:1.4rem;color:{ss_color};">{self_svc}%</span>
        <span style="font-family:DM Mono,monospace;font-size:0.6rem;color:#4B5568;">
          vs avg {b["self_service_avg"]}% · top {b["self_service_top"]}%
        </span>
      </div>
      <div style="display:flex;align-items:center;gap:6px;">
        {pct_bar(self_svc, 100, ss_color)}
        <div style="width:1px;height:8px;background:#3D4D66;position:relative;
             margin-left:{b["self_service_avg"]}%;"></div>
      </div>
    </div>

    <div style="background:#0E1118;border-radius:8px;padding:0.6rem 0.8rem;">
      <div style="font-family:DM Mono,monospace;font-size:0.55rem;color:#4B5568;margin-bottom:4px;">
        MIGRATION COMPLEXITY
      </div>
      <div style="display:flex;align-items:baseline;gap:0.4rem;margin-bottom:6px;">
        <span style="font-family:Syne,sans-serif;font-weight:800;
              font-size:1.4rem;color:{mig_color_b};">{ms}/100</span>
        <span style="font-family:DM Mono,monospace;font-size:0.6rem;color:#4B5568;">
          vs avg {b["migration_avg"]}/100
        </span>
      </div>
      {pct_bar(ms, 100, mig_color_b)}
    </div>
  </div>

  <div style="font-family:Plus Jakarta Sans,sans-serif;font-size:0.72rem;
       color:#3D4D66;line-height:1.5;border-top:1px solid #0E1118;padding-top:0.5rem;">
    {b["note"]}
  </div>
</div>
'''



def render_export_buttons(analysis, flow, results=None, flows_map=None, raw_yamls=None, mode='individual'):
    """
    Botones de exportación homologados: PDF individual + Excel individual + 
    PDF portfolio + Excel portfolio. Misma UI en ambos modos.
    """
    if mode == 'individual':
        excel_key = 'excel_bytes_main'
        pdf_key   = 'pdf_bytes_main'
        if excel_key not in st.session_state: st.session_state[excel_key] = None
        if pdf_key   not in st.session_state: st.session_state[pdf_key]   = None

        st.markdown(
            '<div style="font-family:DM Mono,monospace;font-size:0.6rem;'
            'color:#3D4D66;letter-spacing:0.1em;margin-bottom:0.5rem;">EXPORT</div>',
            unsafe_allow_html=True)
        ecol, pcol = st.columns(2)
        with ecol:
            if st.button('Export Excel', key='btn_excel_main', use_container_width=True):
                with st.spinner('Generating Excel...'):
                    try:
                        fname = (flow.flow_name if flow else 'flow') + '.yaml'
                        single_result = dict(analysis)
                        single_result['filename'] = fname
                        raw_yaml_map = {}
                        st.session_state[excel_key] = generar_portfolio_excel_v2(
                            [single_result], raw_yaml_map)
                    except Exception as e: st.error(str(e))
            if st.session_state.get(excel_key):
                st.download_button('↓ Download Excel', data=st.session_state[excel_key],
                    file_name='orchestria_flow.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    key='dl_excel_main', use_container_width=True)
        with pcol:
            if st.button('Generate PDF', key='btn_pdf_main', use_container_width=True):
                with st.spinner('Building report...'):
                    try: st.session_state[pdf_key] = generar_pdf_bytes(flow, analysis)
                    except Exception as e: st.error(str(e))
            if st.session_state.get(pdf_key):
                fname_dl = (flow.flow_name if flow else 'flow') + '_orchestria.pdf'
                st.download_button('↓ Download PDF', data=st.session_state[pdf_key],
                    file_name=fname_dl, mime='application/pdf',
                    key='dl_pdf_main', use_container_width=True)

    else:  # portfolio
        excel_key = 'excel_bytes_batch'
        pdf_key   = 'pdf_bytes_batch'
        if excel_key not in st.session_state: st.session_state[excel_key] = None
        if pdf_key   not in st.session_state: st.session_state[pdf_key]   = None

        st.markdown(
            '<div style="font-family:DM Mono,monospace;font-size:0.6rem;'
            'color:#3D4D66;letter-spacing:0.1em;margin-bottom:0.5rem;">EXPORT PORTFOLIO</div>',
            unsafe_allow_html=True)
        ecol, pcol = st.columns(2)
        with ecol:
            st.markdown('<div style="font-family:DM Mono,monospace;font-size:0.55rem;color:#4B5568;margin-bottom:3px;">TECHNICAL</div>', unsafe_allow_html=True)
            if st.button('Export Excel', key='btn_excel_batch', use_container_width=True):
                with st.spinner('Generating Excel...'):
                    try: st.session_state[excel_key] = generar_portfolio_excel_v2(
                            results, raw_yamls or {})
                    except Exception as e: st.error(str(e))
            if st.session_state.get(excel_key):
                st.download_button('↓ Download Excel', data=st.session_state[excel_key],
                    file_name='orchestria_portfolio.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    key='dl_excel_batch', use_container_width=True)
        with pcol:
            st.markdown('<div style="font-family:DM Mono,monospace;font-size:0.55rem;color:#4B5568;margin-bottom:3px;">EXECUTIVE</div>', unsafe_allow_html=True)
            if st.button('Generate PDF', key='btn_pdf_batch', use_container_width=True):
                with st.spinner('Generating PDF...'):
                    try: st.session_state[pdf_key] = generar_portfolio_pdf(
                            results, flows_map or {})
                    except Exception as e: st.error(str(e))
            if st.session_state.get(pdf_key):
                st.download_button('↓ Download PDF', data=st.session_state[pdf_key],
                    file_name='orchestria_portfolio_report.pdf',
                    mime='application/pdf',
                    key='dl_pdf_batch', use_container_width=True)


# ── HEADER ─────────────────────────────────────────────────────────────────────
st.markdown(
    '<div class="orch-header">'
    '<div style="display:flex;align-items:center;">'
    '<span class="orch-logo-main">Orchestr</span>'
    '<span class="orch-logo-dot">IA</span>'
    '<span class="orch-logo-badge">IVR · IA</span>'
    '</div>'
    '<span class="orch-tagline">Genesys Flow Intelligence · YAML · JSON · XML</span>'
    '</div>',
    unsafe_allow_html=True)

modo = st.radio('', ['Individual Flow', 'Portfolio Batch'],
                horizontal=True, label_visibility='collapsed')
st.markdown('<div style="margin-top:1.75rem;"></div>', unsafe_allow_html=True)

YAML_EXAMPLE = '''inboundCall:
  name: "Flujo Banca Retail"
  startUpRef: "./menus/menu_principal"
  defaultLanguage: "es-ES"
  supportedLanguages:
    es-ES:
      textToSpeech: Genesys TTS'''

# ── INDIVIDUAL ───────────────────────────────────────────────────────────────
if modo == 'Individual Flow':
    col_l, col_r = st.columns([1, 1], gap='large')

    with col_l:
        st.markdown('<span class="lbl">Upload Flow</span>', unsafe_allow_html=True)
        uploaded = st.file_uploader('', type=['yaml','yml','json','xml'],
                                    label_visibility='collapsed')
        st.markdown(
            '<div style="margin:0.6rem 0;text-align:center;font-family:DM Mono,'
            'monospace;font-size:0.6rem;color:#1E2535;letter-spacing:0.12em;">'
            '— OR PASTE YAML DIRECTLY —</div>',
            unsafe_allow_html=True)
        yaml_input = st.text_area('', height=220, placeholder=YAML_EXAMPLE,
                                  label_visibility='collapsed')
        st.markdown(
            '<div style="display:flex;flex-wrap:wrap;gap:5px;margin-top:0.6rem;">'
            '<span class="compat-chip">✓ inboundCall</span>'
            '<span class="compat-chip">✓ YAML / JSON / XML</span>'
            '<span class="compat-chip">✓ Genesys Cloud</span>'
            '<span class="compat-chip">✓ Architect flows</span>'
            '</div>',
            unsafe_allow_html=True)

    with col_r:
        analizar_clicked = st.button('Analyze Flow →', type='primary', use_container_width=True)
        loading_slot = st.empty()

        # Empty state solo cuando no hay análisis ni input
        if not st.session_state.analysis and not uploaded and not yaml_input.strip():
            st.markdown(empty_state_panel(), unsafe_allow_html=True)

        if analizar_clicked:
            content, filename = '', 'flow.yaml'
            if uploaded:
                content, filename = uploaded.read().decode('utf-8'), uploaded.name
            elif yaml_input.strip():
                content = yaml_input.strip()
            else:
                st.error('Upload a file or paste YAML content.')
                st.stop()

            PHASES = [
                "Parsing YAML structure",
                "Extracting node inventory",
                "Running migration model",
                "Generating AI analysis",
                "Building report",
            ]

            loading_slot.markdown(
                ivr_loading_panel(0, 1, filename, phases=PHASES),
                unsafe_allow_html=True)
            time.sleep(0.3)

            flow, err = parse_content(content, filename)
            if err:
                loading_slot.empty(); st.error(err); st.stop()

            loading_slot.markdown(ivr_loading_panel(1, 1, filename, phases=PHASES), unsafe_allow_html=True)
            time.sleep(0.2)
            loading_slot.markdown(ivr_loading_panel(2, 1, filename, phases=PHASES), unsafe_allow_html=True)
            time.sleep(0.2)
            loading_slot.markdown(ivr_loading_panel(3, 1, filename, phases=PHASES), unsafe_allow_html=True)

            analysis = IVRAnalyzer().analyze(flow)

            loading_slot.markdown(ivr_loading_panel(4, 1, filename, phases=PHASES), unsafe_allow_html=True)
            time.sleep(0.3)
            loading_slot.empty()

            st.session_state.analysis = analysis
            st.session_state.flow = flow
            st.session_state['pdf_bytes_main'] = None
            st.session_state['excel_bytes_main'] = None
            st.rerun()

    if st.session_state.analysis and modo == 'Individual Flow':
        st.divider()
        # Export buttons — homologados con batch
        render_export_buttons(
            st.session_state.analysis,
            st.session_state.flow,
            mode='individual')
        st.divider()
        mostrar_resultado(st.session_state.analysis,
                          flow=st.session_state.flow, key_prefix='main')

# ── BATCH ─────────────────────────────────────────────────────────────────────
else:
    if 'queued_files' not in st.session_state:
        st.session_state.queued_files = {}

    col_l, col_r = st.columns([1, 1], gap='large')

    with col_l:
        st.markdown('<span class="lbl">Upload Portfolio · up to 50 flows</span>',
                    unsafe_allow_html=True)
        new_uploads = st.file_uploader('', type=['yaml','yml','json','xml'],
                                       accept_multiple_files=True,
                                       label_visibility='collapsed',
                                       key='batch_uploader')
        if new_uploads:
            for f in new_uploads:
                if f.name not in st.session_state.queued_files:
                    st.session_state.queued_files[f.name] = f.read()
        queued  = st.session_state.queued_files
        total_q = len(queued)
        if total_q > 0:
            st.markdown(
                f'<div style="font-family:DM Mono,monospace;font-size:0.7rem;'
                f'color:#00D4AA;margin-top:0.5rem;">'
                f'{total_q} file(s) queued · {min(total_q,50)} will be analyzed</div>',
                unsafe_allow_html=True)
            for fname in list(queued.keys()):
                c1, c2 = st.columns([6,1])
                with c1:
                    st.markdown(
                        f'<div style="font-family:DM Mono,monospace;font-size:0.7rem;'
                        f'color:#4B5568;padding:2px 0;">{fname}</div>',
                        unsafe_allow_html=True)
                with c2:
                    if st.button('✕', key=f'del_{fname}', help=f'Remove {fname}'):
                        del st.session_state.queued_files[fname]
                        st.rerun()
            if st.button('Clear all', key='clear_queue'):
                st.session_state.queued_files = {}
                st.rerun()

    with col_r:
        run_batch = st.button('Analyze Portfolio →', type='primary',
                               disabled=not st.session_state.queued_files)
        loading_slot = st.empty()
        if not st.session_state.batch_results and not st.session_state.queued_files:
            st.markdown(empty_state_panel(), unsafe_allow_html=True)

    # Feature chips — fuera de columnas, mismos en ambos modos
    st.markdown(
        '<div style="display:flex;flex-wrap:wrap;gap:5px;margin-top:1rem;">'
        '<span class="compat-chip">✓ Up to 50 flows</span>'
        '<span class="compat-chip">✓ Add files one by one</span>'
        '<span class="compat-chip">✓ Auto-ranked by score</span>'
        '<span class="compat-chip">✓ Excel + PDF export</span>'
        '</div>', unsafe_allow_html=True)

    uploaded_files = st.session_state.queued_files

    if run_batch and uploaded_files:
        resultados, flows_map, raw_yamls_map = [], {}, {}
        total_files = len(uploaded_files)
        analyzer = IVRAnalyzer()
        for i, (fname_k, fbytes_k) in enumerate(list(uploaded_files.items())[:50]):
            loading_slot.markdown(
                ivr_loading_panel(i + 1, total_files, fname_k),
                unsafe_allow_html=True)
            raw_content = fbytes_k.decode('utf-8') if isinstance(fbytes_k, bytes) else fbytes_k
            flow, err = parse_content(raw_content, fname_k)
            if err:
                resultados.append({'filename':fname_k,'score':0,'error':err})
            else:
                r = analyzer.analyze(flow)
                r['filename'] = fname_k
                resultados.append(r)
                flows_map[fname_k] = flow
                try:
                    import yaml as _yaml
                    raw_yamls_map[fname_k] = _yaml.safe_load(raw_content)
                except: pass
        loading_slot.empty()
        st.session_state.batch_results  = resultados
        st.session_state.batch_flows    = flows_map
        st.session_state.batch_raw_yamls = raw_yamls_map
        st.rerun()

    if st.session_state.batch_results:
        results   = st.session_state.batch_results
        flows_map = st.session_state.batch_flows
        ok = sorted([r for r in results if 'error' not in r],
                    key=lambda x: x.get('score',0), reverse=True)

        # Portfolio summary card
        st.markdown(portfolio_summary_card(results), unsafe_allow_html=True)

        # Export buttons — homologados con individual
        render_export_buttons(
            None, None,
            results=results,
            flows_map=flows_map,
            raw_yamls=st.session_state.get('batch_raw_yamls',{}),
            mode='portfolio')

        st.markdown(
            f'<div style="margin:1.5rem 0 1rem;" class="lbl">'
            f'Portfolio · {len(ok)} flows · ranked by quality score</div>',
            unsafe_allow_html=True)

        for idx, r in enumerate(ok):
            s  = r.get('score',0)
            ml = r.get('inventory',{}).get('migration_level','—')
            fname = r['filename'].replace('.yaml','').replace('.yml','')
            dot = '🟢' if s >= 70 else '🟡' if s >= 40 else '🔴'
            with st.expander(f'{dot}  {fname}   ·   {s}/100   ·   {ml}'):
                mostrar_resultado(r, flow=flows_map.get(r['filename']),
                                  key_prefix=f'batch_{idx}')
        for r in [x for x in results if 'error' in x]:
            st.error(f"Error · {r['filename']}: {r['error']}")
